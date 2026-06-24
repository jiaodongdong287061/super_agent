from __future__ import annotations

import logging
from pathlib import Path

from langchain_core.documents import Document

from super_agent.knowledge.loaders.base import BaseLoader

logger = logging.getLogger(__name__)


class ExcelLoader(BaseLoader):
    def __init__(self, chunk_size: int = 20, chunk_overlap: int = 3) -> None:
        if chunk_size <= 0:
            raise ValueError("chunk_size must be positive")
        if chunk_overlap < 0:
            raise ValueError("chunk_overlap must be non-negative")
        if chunk_overlap >= chunk_size:
            raise ValueError("chunk_overlap must be less than chunk_size")
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def load(self, source: str) -> list[Document]:
        ext = Path(source).suffix.lower()
        if ext == ".xlsx":
            return self._load_xlsx(source)
        if ext == ".xls":
            return self._load_xls(source)
        raise ValueError(f"Unsupported extension: {ext}")

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def _load_xlsx(self, source: str) -> list[Document]:
        import openpyxl

        wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
        docs: list[Document] = []
        file_name = Path(source).name

        for sheet in wb.worksheets:
            sheet_docs = self._process_xlsx_sheet(sheet, file_name)
            docs.extend(sheet_docs)

        wb.close()
        return docs

    def _process_xlsx_sheet(self, sheet, file_name: str) -> list[Document]:
        if sheet.max_row is None or sheet.max_row < 1:
            return []

        # 合并单元格填充
        self._fill_merged_cells(sheet)

        # 读取所有行
        all_rows: list[list[str]] = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            all_rows.append(cells)

        if not all_rows:
            return []

        # 过滤空列：找出所有行中至少有一个非空值的列
        max_cols = max(len(r) for r in all_rows) if all_rows else 0
        non_empty_cols: list[int] = []
        for col_idx in range(max_cols):
            for row in all_rows:
                if col_idx < len(row) and row[col_idx].strip():
                    non_empty_cols.append(col_idx)
                    break

        if not non_empty_cols:
            return []

        # 只保留非空列
        filtered_rows: list[list[str]] = []
        for row in all_rows:
            filtered_rows.append([row[i] if i < len(row) else "" for i in non_empty_cols])

        # 表头 = 第一个非空行
        header_idx = 0
        for i, row in enumerate(filtered_rows):
            if any(cell.strip() for cell in row):
                header_idx = i
                break

        headers = filtered_rows[header_idx]
        data_rows = filtered_rows[header_idx + 1 :]

        # 过滤全空数据行
        data_rows = [r for r in data_rows if any(cell.strip() for cell in r)]

        if not data_rows:
            return []

        sheet_name = sheet.title
        return self._chunk_sheet(headers, data_rows, file_name, sheet_name)

    def _fill_merged_cells(self, sheet) -> None:
        """将合并单元格的值向下填充到所有被合并的行，然后取消合并。"""
        if not sheet.merged_cells.ranges:
            return
        ranges = list(sheet.merged_cells.ranges)
        for mr in ranges:
            top_value = sheet.cell(row=mr.min_row, column=mr.min_col).value
            if top_value is None:
                top_value = ""
            # 先取消合并（使其可写），再填充值
            sheet.unmerge_cells(str(mr))
            for row_idx in range(mr.min_row, mr.max_row + 1):
                for col_idx in range(mr.min_col, mr.max_col + 1):
                    sheet.cell(row=row_idx, column=col_idx, value=top_value)

    def _load_xls(self, source: str) -> list[Document]:
        import xlrd

        wb = xlrd.open_workbook(source)
        docs: list[Document] = []
        file_name = Path(source).name

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_docs = self._process_xls_sheet(sheet, file_name)
            docs.extend(sheet_docs)

        return docs

    def _process_xls_sheet(self, sheet, file_name: str) -> list[Document]:
        if sheet.nrows < 2:
            return []

        all_rows: list[list[str]] = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            all_rows.append(cells)

        # 过滤空列
        max_cols = max(len(r) for r in all_rows) if all_rows else 0
        non_empty_cols: list[int] = []
        for col_idx in range(max_cols):
            for row in all_rows:
                if col_idx < len(row) and row[col_idx].strip():
                    non_empty_cols.append(col_idx)
                    break

        if not non_empty_cols:
            return []

        filtered_rows: list[list[str]] = []
        for row in all_rows:
            filtered_rows.append([row[i] if i < len(row) else "" for i in non_empty_cols])

        header_idx = 0
        for i, row in enumerate(filtered_rows):
            if any(cell.strip() for cell in row):
                header_idx = i
                break

        headers = filtered_rows[header_idx]
        data_rows = filtered_rows[header_idx + 1 :]
        data_rows = [r for r in data_rows if any(cell.strip() for cell in r)]

        if not data_rows:
            return []

        sheet_name = sheet.name
        return self._chunk_sheet(headers, data_rows, file_name, sheet_name)

    def _chunk_sheet(
        self,
        headers: list[str],
        data_rows: list[list[str]],
        file_name: str,
        sheet_name: str,
    ) -> list[Document]:
        docs: list[Document] = []
        total = len(data_rows)
        step = self.chunk_size - self.chunk_overlap
        if step <= 0:
            step = 1

        pos = 0
        while pos < total:
            end = min(pos + self.chunk_size, total)
            chunk_rows = data_rows[pos:end]
            row_start = pos + 1
            row_end = end

            header_text = self._format_header(headers)
            data_text = self._format_data_rows(headers, chunk_rows, row_start, row_end)
            page_content = f"{header_text}\n\n{data_text}"

            meta = {
                "source": file_name,
                "sheet_name": sheet_name,
                "row_range": f"{row_start}-{row_end}",
                "headers": headers,
            }
            docs.append(Document(page_content=page_content, metadata=meta))
            pos += step

        return docs

    @staticmethod
    def _format_header(headers: list[str]) -> str:
        pairs = [f"{h}: {h}" for h in headers]
        return f"[表头]\n{', '.join(pairs)}"

    @staticmethod
    def _format_data_rows(
        headers: list[str],
        rows: list[list[str]],
        row_start: int,
        row_end: int,
    ) -> str:
        lines = [f"[数据行 {row_start}-{row_end}]"]
        for row in rows:
            pairs: list[str] = []
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else ""
                if val.strip():
                    pairs.append(f"{h}: {val}")
            lines.append(", ".join(pairs))
        return "\n".join(lines)
