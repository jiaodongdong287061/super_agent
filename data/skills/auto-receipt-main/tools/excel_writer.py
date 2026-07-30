"""
Excel 写入工具

功能：将结构化数据写入报销明细 Excel
"""

from typing import List, Dict, Any
from pathlib import Path
from openpyxl import Workbook


def write_reimbursement_excel(
    data: List[Dict[str, Any]],
    output_path: str,
    company_name: str = None
) -> str:
    """
    生成报销明细 Excel 文件

    Args:
        data: 报销数据列表，每项为字典
              必需字段：起-年月日, 止-年月日, 起-地点, 止-地点, 费用种类, 单据张数, 金额
        output_path: 输出文件路径
        company_name: 公司名称（可选，用于日志）

    Returns:
        保存的文件路径

    Raises:
        ValueError: 数据格式错误
    """
    if not data:
        raise ValueError("数据不能为空")

    # 验证数据格式
    required_fields = ['起-年月日', '止-年月日', '起-地点', '止-地点', '费用种类', '单据张数', '金额']
    for i, row in enumerate(data):
        for field in required_fields:
            if field not in row:
                raise ValueError(f"第 {i+1} 行缺少必需字段: {field}")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active

    # 第1行：标题
    ws.cell(row=1, column=1, value="报销明细")

    # 第2行：表头
    headers = ['起-年月日', '止-年月日', '起-地点', '止-地点', '费用种类', '单据张数', '金额']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header)

    # 第3行起：数据
    for row_idx, row_data in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1, value=row_data.get('起-年月日', ''))
        ws.cell(row=row_idx, column=2, value=row_data.get('止-年月日', ''))
        ws.cell(row=row_idx, column=3, value=row_data.get('起-地点', ''))
        ws.cell(row=row_idx, column=4, value=row_data.get('止-地点', ''))
        ws.cell(row=row_idx, column=5, value=row_data.get('费用种类', ''))
        ws.cell(row=row_idx, column=6, value=row_data.get('单据张数', 0))
        ws.cell(row=row_idx, column=7, value=row_data.get('金额', 0))

    # 保存
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    return str(output_file)


def create_sample_data() -> List[Dict[str, Any]]:
    """
    创建示例数据，用于测试
    """
    return [
        {
            "起-年月日": "2025-02-26",
            "止-年月日": "2025-02-26",
            "起-地点": "上海",
            "止-地点": "北京",
            "费用种类": "机票",
            "单据张数": 1,
            "金额": 640.00
        },
        {
            "起-年月日": "2025-02-26",
            "止-年月日": "2025-02-26",
            "起-地点": "上海",
            "止-地点": "上海",
            "费用种类": "住宿费",
            "单据张数": 1,
            "金额": 957.00
        },
        {
            "起-年月日": "2025-02-26",
            "止-年月日": "2025-02-26",
            "起-地点": "上海",
            "止-地点": "上海",
            "费用种类": "打车费",
            "单据张数": 1,
            "金额": 78.95
        },
        {
            "起-年月日": "2025-02-26",
            "止-年月日": "2025-02-26",
            "起-地点": "北京",
            "止-地点": "北京",
            "费用种类": "餐补费",
            "单据张数": 2,
            "金额": 240.00
        }
    ]


if __name__ == "__main__":
    # 测试
    data = create_sample_data()
    output = write_reimbursement_excel(data, "./output/报销明细_输出.xlsx")
    print(f"已生成: {output}")
