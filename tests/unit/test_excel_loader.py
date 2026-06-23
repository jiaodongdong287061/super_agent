import pytest
from pathlib import Path
from openpyxl import Workbook

from super_agent.knowledge.loaders.excel import ExcelLoader


def _make_xlsx(tmp_path: Path, rows: list[list[str]], headers: list[str], sheet_name: str = "Sheet") -> Path:
    """辅助函数：用 openpyxl 创建一个简单 xlsx 文件并返回路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    p = tmp_path / "test.xlsx"
    wb.save(str(p))
    return p


class TestExcelLoaderXlsxSingleSheet:
    def test_load_returns_documents(self, tmp_path):
        headers = ["姓名", "部门", "级别"]
        rows = [
            ["张三", "SRE", "P7"],
            ["李四", "DBA", "P6"],
        ]
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader()
        docs = loader.load(str(p))
        assert len(docs) >= 1

    def test_document_contains_header_and_data(self, tmp_path):
        headers = ["姓名", "部门"]
        rows = [["张三", "SRE"]]
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader()
        docs = loader.load(str(p))
        assert len(docs) == 1
        text = docs[0].page_content
        assert "[表头]" in text
        assert "姓名: 姓名" in text
        assert "部门: 部门" in text
        assert "姓名: 张三" in text
        assert "部门: SRE" in text

    def test_metadata_fields(self, tmp_path):
        headers = ["姓名", "部门"]
        rows = [["张三", "SRE"]]
        p = _make_xlsx(tmp_path, rows, headers, sheet_name="员工表")
        loader = ExcelLoader()
        docs = loader.load(str(p))
        meta = docs[0].metadata
        assert meta["source"] == "test.xlsx"
        assert meta["sheet_name"] == "员工表"
        assert meta["headers"] == ["姓名", "部门"]
        assert "row_range" in meta

    def test_supported_extensions(self):
        loader = ExcelLoader()
        assert ".xlsx" in loader.supported_extensions()
        assert ".xls" in loader.supported_extensions()


class TestExcelLoaderMergedCells:
    def test_merged_cell_fill_down(self, tmp_path):
        """合并单元格：左上角的值应填充到所有被合并的行。"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "合并测试"
        ws.append(["系统", "指标", "值"])
        # 合并 A2:A4（系统列前 3 行数据）
        ws.append(["生产环境", "CPU", "80%"])
        ws.append([None, "内存", "60%"])
        ws.append([None, "磁盘", "50%"])
        ws.append(["测试环境", "CPU", "30%"])
        ws.merge_cells("A2:A4")

        p = tmp_path / "merged.xlsx"
        wb.save(str(p))

        loader = ExcelLoader()
        docs = loader.load(str(p))
        text = docs[0].page_content
        # 合并填充后，A2-A4 都应为 "生产环境"
        assert text.count("系统: 生产环境") == 3
        assert "系统: 测试环境" in text

    def test_no_merge_in_xls(self, tmp_path):
        """xls 不支持合并填充，空值保留为空。"""
        pass


class TestExcelLoaderChunking:
    def test_chunk_size_splits_rows(self, tmp_path):
        """超过 chunk_size 的数据行应拆分为多个 Document。"""
        headers = ["ID", "名称"]
        rows = [[str(i), f"项目{i}"] for i in range(1, 26)]  # 25 行数据
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader(chunk_size=10, chunk_overlap=0)
        docs = loader.load(str(p))
        assert len(docs) == 3  # 10 + 10 + 5
        assert "1-10" in docs[0].metadata["row_range"]
        assert "11-20" in docs[1].metadata["row_range"]
        assert "21-25" in docs[2].metadata["row_range"]

    def test_chunk_overlap(self, tmp_path):
        """相邻 chunk 应有重叠行。"""
        headers = ["ID", "名称"]
        rows = [[str(i), f"项目{i}"] for i in range(1, 26)]  # 25 行数据
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader(chunk_size=10, chunk_overlap=3)
        docs = loader.load(str(p))
        # step = 10 - 3 = 7; chunks: [0:10], [7:17], [14:24], [21:25]
        assert len(docs) == 4
        # 验证重叠：chunk1 第一个数据行应包含 chunk0 倒数第3行的数据
        assert "8" in docs[1].page_content  # row 8 出现在 chunk1

    def test_small_data_single_chunk(self, tmp_path):
        """数据行不足 chunk_size 时生成单个 chunk。"""
        headers = ["ID"]
        rows = [["1"], ["2"]]
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader(chunk_size=20, chunk_overlap=3)
        docs = loader.load(str(p))
        assert len(docs) == 1
        assert "1-2" in docs[0].metadata["row_range"]

    def test_each_chunk_contains_header(self, tmp_path):
        """每个 chunk 都应包含表头部分。"""
        headers = ["ID", "名称"]
        rows = [[str(i), f"项目{i}"] for i in range(1, 26)]
        p = _make_xlsx(tmp_path, rows, headers)
        loader = ExcelLoader(chunk_size=10, chunk_overlap=0)
        docs = loader.load(str(p))
        for doc in docs:
            assert "[表头]" in doc.page_content
